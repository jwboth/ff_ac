"""Generate pressure_temperature_protocol.csv per AC experiment.

The mass computation in DarSIA reads the pressure/temperature *state* at the
experiment start to set CO2 gas density and aqueous solubility. Two sources:

* ``--pressure-xlsx``: real barometric data from the Florida (Bergen) weather
  station (columns ``Dato, Tid, Lufttrykk`` in **hPa**, 10-min resolution). For
  each experiment we take the AVERAGE pressure from the experiment start to
  +``--window-hours`` (default 36 h), apply an altitude correction (the station
  is ~40 m HIGHER than the rig, so the rig pressure is slightly higher), and
  convert hPa -> bar.
* fallback constant: ``--pressure-bar`` / ``--temperature-c`` if no xlsx (or if
  an experiment's window is not covered by the data).

Altitude correction (hypsometric):  P_rig = P_station * exp(g*dh/(R*T)),
with dh = ``--altitude-diff-m`` (rig metres BELOW the station), R=287.05,
T = temperature in Kelvin. 40 m -> ~ +0.46 % (~ +4.7 hPa).

Usage
-----
    python scripts/generate_pressure.py --protocols-root protocols --all \
        --pressure-xlsx "../*Florida*.xlsx" --altitude-diff-m 40 --window-hours 36
"""

from __future__ import annotations

import argparse
import csv
import glob
import logging
import math
from datetime import datetime, timedelta
from pathlib import Path

LOG = logging.getLogger("generate_pressure")

R_AIR = 287.05      # J/(kg K) specific gas constant for dry air
G = 9.80665         # m/s^2


def load_station_series(patterns: list) -> list:
    """Load (datetime, hPa) from one or more Florida xlsx files; merged + sorted."""
    import openpyxl
    series: dict = {}
    files = []
    for pat in patterns:
        files += glob.glob(pat)
    for f in sorted(set(files)):
        wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        it = ws.iter_rows(values_only=True)
        header = [str(h).strip().lower() if h else "" for h in next(it)]
        try:
            i_date, i_time, i_p = (header.index("dato"), header.index("tid"),
                                   header.index("lufttrykk"))
        except ValueError:
            LOG.warning("%s: unexpected header %s; skipping", f, header)
            wb.close()
            continue
        dropped = 0
        for row in it:
            if row[i_date] is None or row[i_p] is None:
                continue
            try:
                p = float(row[i_p])
            except (TypeError, ValueError):
                continue
            # Drop sentinel / error readings (the station uses 9999.99 for
            # "no data"). Keep only physically plausible sea-level pressures;
            # the all-time world records are ~870 and ~1084 hPa.
            if not (870.0 <= p <= 1085.0):
                dropped += 1
                continue
            dt = _parse_dt(row[i_date], row[i_time])
            if dt is not None:
                series[dt] = p
        if dropped:
            LOG.info("%s: dropped %d implausible/sentinel pressure rows",
                     Path(f).name, dropped)
        wb.close()
        LOG.info("Loaded %s (%d total points)", Path(f).name, len(series))
    return sorted(series.items())


def _parse_dt(date_val, time_val):
    d = str(date_val).strip()[:10]
    t = str(time_val).strip() if time_val is not None else "00:00"
    for fmt in ("%Y-%m-%d %H:%M", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(f"{d} {t}", fmt)
        except ValueError:
            continue
    try:
        return datetime.strptime(d, "%Y-%m-%d")
    except ValueError:
        return None


def altitude_factor(altitude_diff_m: float, temperature_c: float) -> float:
    """P_rig / P_station; rig is altitude_diff_m BELOW the station."""
    return math.exp(G * altitude_diff_m / (R_AIR * (temperature_c + 273.15)))


def _experiment_start(protocol_dir: Path):
    inj = protocol_dir / "injection_protocol.csv"
    if not inj.exists():
        return None
    try:
        rows = list(csv.DictReader(inj.open(encoding="utf-8")))
        return datetime.fromisoformat(rows[0]["start"]) if rows else None
    except (KeyError, ValueError, StopIteration):
        return None


def average_pressure_hpa(series, start, window_hours):
    end = start + timedelta(hours=window_hours)
    vals = [p for (dt, p) in series if start <= dt <= end]
    return (sum(vals) / len(vals)) if vals else None


def write_pressure_csv(protocol_dir, pressure_bar, temperature_c, span_days=60):
    start = _experiment_start(protocol_dir)
    if start is None:
        LOG.warning("%s: no injection_protocol.csv start; skipping", protocol_dir.name)
        return False
    out = protocol_dir / "pressure_temperature_protocol.csv"
    with out.open("w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(["datetime", "pressure_bar", "temperature_celsius",
                    "pressure_gradient_bar", "temperature_gradient_celsius"])
        for t in (start - timedelta(days=1), start + timedelta(days=span_days)):
            w.writerow([t.isoformat(sep=" "), round(pressure_bar, 6), temperature_c, 0.0, 0.0])
    return True


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--protocols-root", required=True, type=Path)
    ap.add_argument("--experiments", nargs="*", default=[])
    ap.add_argument("--all", action="store_true")
    ap.add_argument("--pressure-xlsx", nargs="*", default=[],
                    help="Florida station xlsx file(s)/glob(s) (hPa)")
    ap.add_argument("--window-hours", type=float, default=36.0)
    ap.add_argument("--altitude-diff-m", type=float, default=40.0,
                    help="rig metres BELOW the station (station is higher)")
    ap.add_argument("--pressure-bar", type=float, default=1.013, help="constant fallback")
    ap.add_argument("--temperature-c", type=float, default=23.0)
    args = ap.parse_args()
    logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")

    root = args.protocols_root
    if args.all:
        dirs = sorted([d for d in root.iterdir()
                       if d.is_dir() and d.name.lower().startswith("ac")],
                      key=lambda d: int("".join(c for c in d.name if c.isdigit()) or 0))
    else:
        dirs = [root / e for e in args.experiments]

    series = load_station_series(args.pressure_xlsx) if args.pressure_xlsx else []
    factor = altitude_factor(args.altitude_diff_m, args.temperature_c)
    if series:
        LOG.info("Station data: %s .. %s | altitude factor x%.5f",
                 series[0][0], series[-1][0], factor)

    n_real = n_const = 0
    for d in dirs:
        if not d.exists():
            continue
        pressure_bar = args.pressure_bar
        if series:
            start = _experiment_start(d)
            avg_hpa = average_pressure_hpa(series, start, args.window_hours) if start else None
            if avg_hpa is not None:
                pressure_bar = (avg_hpa * factor) / 1000.0  # hPa -> bar
                LOG.info("%s: avg %.1f hPa (+alt) -> %.5f bar", d.name, avg_hpa, pressure_bar)
                n_real += 1
            else:
                LOG.warning("%s: window not covered by station data; using constant %.3f bar",
                            d.name, args.pressure_bar)
                n_const += 1
        else:
            n_const += 1
        write_pressure_csv(d, pressure_bar, args.temperature_c)
    LOG.info("Done: %d from station data, %d from constant fallback", n_real, n_const)


if __name__ == "__main__":
    main()
