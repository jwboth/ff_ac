"""Distributed cross-run Wasserstein queue for ff_ac.

Ports the ff_um ``distributed_wasserstein_queue.py`` command surface onto the new
DarSIA Wasserstein computation. Enumerates run pairs C(n,2), lays a shared time
grid over each pair (so every experiment is compared at the *same*
time-since-injection), and queues one task per (pair, time-point, ROI).

A built-in budget check compares the planned number of comparisons against the
ff_um reference (108 137 matched image-pairs / 306 726 ROI-distances) and warns
if the plan exceeds it - the constraint that drives image sub-selection for the
43-experiment AC series.

Commands::

    python scripts/distributed_wasserstein_queue.py supervisor \
        --queue Z:\\FF_AC\\Analysis\\Wasserstein \
        --runs 17 18 ... 61 --points-per-run 110 --rois full box1 box2 \
        --resize 0.10 --tasks-per-file 40 --prepare-queue --reset-queue [--dry-run]

    python scripts/distributed_wasserstein_queue.py watchdog \
        --queue ... --workers 16 --control-dir ...\\worker_control --stop-when-drained

    python scripts/distributed_wasserstein_queue.py assemble \
        --queue ... --out-path analysis/wasserstein_pairs_all.xlsx

The W1 computation per task is delegated to a configurable ``--worker-cmd`` (a
DarSIA comparison call); confirm its interface before a production run.
"""

from __future__ import annotations

import argparse
import logging
import sys
from itertools import combinations
from pathlib import Path

import queue_core as qc
import _workflow_common as wc

LOG = logging.getLogger("wasserstein_queue")

def _last_float(text):
    for line in reversed((text or "").splitlines()):
        line = line.strip()
        if not line:
            continue
        try:
            return float(line)
        except ValueError:
            continue
    return None


UM_MATCHED_PAIRS = 108_137
UM_ROI_DISTANCES = 306_726

WASS_WORKER_CMD = (sys.executable + " scripts/wasserstein_worker.py "
                   "--run-a {run_a} --run-b {run_b} --time-index {time_index} "
                   "--roi {roi} --resize {resize}")


def _pairs(runs: list, explicit) -> list:
    if explicit:
        return [tuple(s.split("-", 1)) for s in explicit]
    return list(combinations(runs, 2))


def _budget_report(n_pairs: int, points: int, rois: int):
    matched = n_pairs * points
    distances = matched * rois
    LOG.info("Plan: %d pairs x %d points = %d matched pairs; x %d ROI = %d distances",
             n_pairs, points, matched, rois, distances)
    LOG.info("ff_um reference: %d matched pairs / %d ROI-distances",
             UM_MATCHED_PAIRS, UM_ROI_DISTANCES)
    if matched > UM_MATCHED_PAIRS or distances > UM_ROI_DISTANCES:
        cap = min(UM_MATCHED_PAIRS // n_pairs, UM_ROI_DISTANCES // (n_pairs * rois))
        LOG.warning("OVER BUDGET vs ff_um - reduce --points-per-run or --rois "
                    "(max points/run = %d for %d pairs, %d ROI)", cap, n_pairs, rois)
    else:
        LOG.info("Within ff_um budget (%.1f%% of matched-pair cap)",
                 100 * matched / UM_MATCHED_PAIRS)
    return matched, distances


def cmd_supervisor(a) -> None:
    pairs = _pairs(a.runs, a.pairs)
    _budget_report(len(pairs), a.points_per_run, len(a.rois))
    if a.dry_run:
        LOG.info("--dry-run: not enqueueing")
        return
    q = qc.Queue(a.queue)
    if a.reset_queue:
        q.reset()
    if a.prepare_queue:
        q.ensure_layout()

    def gen_tasks():
        for (ra, rb) in pairs:
            for ti in range(a.points_per_run):
                for roi in a.rois:
                    yield {"kind": "wasserstein", "run_a": ra, "run_b": rb,
                           "time_index": ti, "roi": roi, "resize": a.resize}

    n = q.enqueue(gen_tasks(), tasks_per_file=a.tasks_per_file)
    LOG.info("Enqueued %d batch(es) for %d pairs", n, len(pairs))


def _make_execute(worker_cmd: str):
    def execute(task: dict) -> dict:
        res = wc.run_subprocess_task(worker_cmd, {
            "run_a": task["run_a"], "run_b": task["run_b"],
            "time_index": task["time_index"], "roi": task["roi"],
            "resize": task["resize"]})
        res["distance"] = _last_float(res.get("stdout", ""))
        res.update({"pair": f"{task['run_a']}-{task['run_b']}",
                    "time_index": task["time_index"], "roi": task["roi"]})
        return res
    return execute


def cmd_worker(a) -> None:
    q = qc.Queue(a.queue)
    qc.run_worker(q, _make_execute(a.worker_cmd), wc.worker_config_from_args(a))


def cmd_watchdog(a) -> None:
    q = qc.Queue(a.queue)
    extra = ["--worker-cmd", a.worker_cmd]
    if a.stop_when_drained:
        extra += ["--exit-when-idle", "--idle-seconds", "3", "--max-tasks", "100000"]
    worker_cmd = wc.build_worker_cmd(__file__, a.queue, extra)
    qc.run_watchdog(q, worker_cmd, wc.watchdog_config_from_args(a),
                    stop_when_drained=a.stop_when_drained)


def cmd_assemble(a) -> None:
    """Assemble results into a pairs workbook matching wasserstein_pairs_all.xlsx."""
    import openpyxl
    q = qc.Queue(a.queue)
    data: dict = {}
    rois: list = []
    for r in qc.iter_results(q):
        pair = r.get("pair")
        if pair is None:
            continue
        roi = r.get("roi")
        if roi not in rois:
            rois.append(roi)
        data.setdefault(pair, {}).setdefault(r.get("time_index"), {})[roi] = r.get("distance")
    out = Path(a.out_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    summary = wb.create_sheet("_summary")
    summary.append(["pair", "rows", "rois"])
    for pair in sorted(data):
        ws = wb.create_sheet(pair[:31])
        ws.append(["time_index", *rois])
        for ti in sorted(data[pair]):
            ws.append([ti, *[data[pair][ti].get(roi) for roi in rois]])
        summary.append([pair, len(data[pair]), ",".join(rois)])
    wb.save(out)
    LOG.info("Assembled %d pair sheet(s) -> %s", len(data), out)
    if not data and not a.allow_missing:
        sys.exit("No results found (use --allow-missing to ignore)")



def cmd_compute(a) -> None:
    """Run the REAL DarSIA cross-run W1 compute (comparison.py), sequential.

    The new DarSIA comparison CLI does not expose skip_existing, so fine-grained
    multi-worker parallelism needs either that one-line flag upstream or the
    queue/worker path. Here we run the genuine compute on one process.
    """
    import subprocess
    cmd = [sys.executable, "scripts/comparison.py", "--config", a.config,
           "--wasserstein-compute"]
    LOG.info("Running: %s", " ".join(cmd))
    if a.workers > 1:
        LOG.warning("--workers>1 requires skip_existing upstream; running 1 process")
    raise SystemExit(subprocess.run(cmd, cwd=str(wc.REPO_ROOT)).returncode)


def cmd_assemble_darsia(a) -> None:
    """Run the REAL DarSIA assemble (comparison.py --wasserstein-assemble)."""
    import subprocess
    cmd = [sys.executable, "scripts/comparison.py", "--config", a.config,
           "--wasserstein-assemble"]
    LOG.info("Running: %s", " ".join(cmd))
    raise SystemExit(subprocess.run(cmd, cwd=str(wc.REPO_ROOT)).returncode)


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    sub = ap.add_subparsers(dest="mode", required=True)

    s = sub.add_parser("supervisor")
    s.add_argument("--queue", required=True)
    s.add_argument("--config", default=None)
    s.add_argument("--runs", nargs="*", default=[])
    s.add_argument("--pairs", nargs="*", default=None, help="explicit a-b pairs")
    s.add_argument("--points-per-run", type=int, default=110)
    s.add_argument("--rois", nargs="+", default=["full", "box1", "box2"])
    s.add_argument("--resize", type=float, default=0.10)
    s.add_argument("--tasks-per-file", type=int, default=40)
    s.add_argument("--prepare-queue", action="store_true")
    s.add_argument("--reset-queue", action="store_true")
    s.add_argument("--dry-run", action="store_true")
    s.set_defaults(func=cmd_supervisor)

    w = sub.add_parser("watchdog")
    wc.add_watchdog_args(w)
    w.add_argument("--worker-cmd", default=WASS_WORKER_CMD)
    w.set_defaults(func=cmd_watchdog)

    wk = sub.add_parser("worker")
    wc.add_worker_args(wk)
    wk.add_argument("--worker-cmd", default=WASS_WORKER_CMD)
    wk.set_defaults(func=cmd_worker)

    g = sub.add_parser("assemble")
    g.add_argument("--queue", required=True)
    g.add_argument("--output", default="xlsx", choices=["xlsx"])
    g.add_argument("--out-path", default="analysis/wasserstein_pairs_all.xlsx")
    g.add_argument("--allow-missing", action="store_true")
    g.set_defaults(func=cmd_assemble)

    cp = sub.add_parser("compute", help="real DarSIA W1 compute via comparison.py")
    cp.add_argument("--config", default="config/wasserstein_ac.toml")
    cp.add_argument("--workers", type=int, default=1)
    cp.set_defaults(func=cmd_compute)

    ad = sub.add_parser("assemble-darsia", help="real DarSIA W1 assemble")
    ad.add_argument("--config", default="config/wasserstein_ac.toml")
    ad.set_defaults(func=cmd_assemble_darsia)

    a = ap.parse_args()
    logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")
    a.func(a)


if __name__ == "__main__":
    main()
