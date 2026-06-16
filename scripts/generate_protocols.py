"""Generate ff_ac protocol CSVs for the Albus (AC) experiment series.

For every experiment this produces the protocol files ff_ac consumes:

* ``imaging_protocol[_<phase>][k].csv``  -> ``path, image_id, datetime``
* ``injection_protocol.csv``             -> ``id, location_x, location_y, start,
                                             end, flow_percent, rate_sccm, density``

The imaging protocol is PURE protocol-based - NO image/EXIF reads. Each image's
time is ``interval_start + (DSC - base) * interval`` where ``interval_start`` is
the folder-name time (identical to the protocol's "Start interval imaging" time,
in the lab/reference clock). Images are SORTED BY TIME (robust to camera-card /
DSC-number resets). A phase split across several folders, and naming variants
(``rest_5min`` vs ``resting_5min``), are handled. Imaging and injection therefore
share the same reference clock, so the analysis time-grid stays aligned.

The injection protocol is reconstructed from the per-experiment
``Albus protocol NN.xlsx`` "Scripted steps" table: the gas-injection ``t=0`` row
anchors the timeline, the Flow % column gives the MFC set-point ramp, and the
"Open L1/L2" notes track the active port.

Usage
-----
    python scripts/generate_protocols.py --albus-root /path/to/Albus \
        --out protocols --experiments AC53      # one experiment
    python scripts/generate_protocols.py --albus-root /path/to/Albus \
        --out protocols --all                    # every experiment
"""

from __future__ import annotations

import argparse
import csv
import logging
import re
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional

LOG = logging.getLogger("generate_protocols")

# --- physical constants (confirmed against the article) ------------------
NOMINAL_SCCM = 10.0      # MFC full scale [ml/min] used as the %-denominator
                         # (confirmed article p3; actual inj. rate <= 2.0 ml/min)
DENSITY_KG_M3 = 1.872    # CO2 std density @ 20 C/1013 mbar (article p4)
PORT_LOCATION = {        # (x, y) in metres
    "L1": (0.45, 0.03),  # port 1 / I1
    "L2": (0.72, 0.18),  # port 2 / I2
}

FOLDER_RE = re.compile(r"(?P<date>\d{6})_(?P<time>\d{6})_(?:AC(?P<num>\d+)_)?(?P<phase>\w+)")
INTERVAL_RE = re.compile(r"(\d+)\s*(s|min)", re.IGNORECASE)
DSC_RE = re.compile(r"DSC0*(\d+)", re.IGNORECASE)


# =========================================================================
# Imaging protocol
# =========================================================================
def _folder_start_and_interval(folder: Path):
    m = FOLDER_RE.search(folder.name)
    start = None
    interval = timedelta(seconds=30)
    if m:
        try:
            start = datetime.strptime(m["date"] + m["time"], "%y%m%d%H%M%S")
        except ValueError:
            LOG.warning("%s: malformed date/time in folder name; using xlsx date",
                        folder.name)
            start = None
    im = INTERVAL_RE.search(folder.name)
    if im:
        n, unit = int(im[1]), im[2].lower()
        interval = timedelta(seconds=n) if unit == "s" else timedelta(minutes=n)
    return start, interval



def build_imaging_protocol(folder: Path, out_csv: Path,
                           experiment_date: Optional[datetime] = None) -> int:
    """Write an imaging protocol for one phase folder. Returns image count.

    PURE protocol-based, no image reads: the interval start time comes from the
    folder name (identical to the protocol's "Start interval imaging" time, in the
    reference/lab clock), and each image's time is ``start + (DSC - base) * interval``.
    """
    jpgs = sorted(folder.glob("*.JPG")) + sorted(folder.glob("*.jpg"))
    jpgs = sorted(set(jpgs), key=lambda p: p.name)
    if not jpgs:
        LOG.warning("No JPGs in %s", folder)
        return 0
    folder_start, interval = _folder_start_and_interval(folder)
    if folder_start is None and experiment_date is not None:
        m = FOLDER_RE.search(folder.name)
        if m:
            try:
                t = datetime.strptime(m["time"], "%H%M%S").time()
                folder_start = datetime.combine(experiment_date.date(), t)
            except ValueError:
                folder_start = None
    if folder_start is None:
        LOG.warning("%s: cannot resolve interval start time; datetimes left blank",
                    folder.name)

    dscs = [int(m[1]) if (m := DSC_RE.search(p.stem)) else None for p in jpgs]
    valid = [d for d in dscs if d is not None]
    if valid and (max(valid) - min(valid) + 1) > len(valid) * 1.5:
        LOG.warning("%s: non-contiguous DSC numbers (card reset?) - using ordinal "
                    "spacing from the interval start", folder.name)

    # Order images by DSC number (== capture order within an interval; for the
    # rare card-reset folder this is the best protocol-only ordering), then place
    # the i-th image at start + i*interval. For contiguous folders this equals
    # start + (DSC - first)*interval exactly; it can never run before the start.
    order = sorted(range(len(jpgs)),
                   key=lambda i: (dscs[i] is None, dscs[i] if dscs[i] is not None else 0,
                                  jpgs[i].name))
    rows = []
    for rank, i in enumerate(order):
        p, d = jpgs[i], dscs[i]
        dt = folder_start + rank * interval if folder_start is not None else None
        rows.append((p.name, d if d is not None else 0, dt))
    out_csv.parent.mkdir(parents=True, exist_ok=True)
    with out_csv.open("w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(["path", "image_id", "datetime"])
        for name, iid, dt in rows:
            w.writerow([name, iid, dt.isoformat(sep=" ") if dt else ""])
    LOG.info("Wrote %s (%d images)", out_csv, len(rows))
    return len(rows)


# =========================================================================
# Injection protocol (from the "Scripted steps" table)
# =========================================================================
def _parse_hms(val):
    if val is None:
        return None
    s = str(val).strip()
    m = re.match(r"(?:(\d+) day[s]?, )?(\d{1,2}):(\d{2}):(\d{2})", s)
    if not m:
        return None
    days = int(m[1]) if m[1] else 0
    return timedelta(days=days, hours=int(m[2]), minutes=int(m[3]), seconds=int(m[4]))


def build_injection_protocol(xlsx: Path, injection_folder: Optional[Path],
                             out_csv: Path,
                             nominal_sccm: float = NOMINAL_SCCM,
                             experiment_date: Optional[datetime] = None) -> int:
    """Reconstruct the injection protocol from the scripted-steps table."""
    import openpyxl

    wb = openpyxl.load_workbook(xlsx, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header_row = None
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            if str(ws.cell(r, c).value).strip().lower() == "target":
                header_row = r
                break
        if header_row:
            break
    if header_row is None:
        LOG.warning("No scripted-steps table in %s; writing placeholder", xlsx.name)
        return _write_placeholder_injection(out_csv)

    COL = {"target": 3, "actual": 5, "todo": 7, "dsc": 11,
           "flow": 13, "what": 15, "seconds": 19, "cum": 20}

    steps = []
    for r in range(header_row + 1, ws.max_row + 1):
        flow = ws.cell(r, COL["flow"]).value
        target = ws.cell(r, COL["target"]).value
        if flow is None and target is None:
            continue
        steps.append({
            "row": r,
            "target": ws.cell(r, COL["target"]).value,
            "actual": ws.cell(r, COL["actual"]).value,
            "todo": str(ws.cell(r, COL["todo"]).value or ""),
            "dsc": ws.cell(r, COL["dsc"]).value,
            "flow": flow,
            "what": str(ws.cell(r, COL["what"]).value or ""),
            "seconds": ws.cell(r, COL["seconds"]).value,
            "cum": _parse_hms(ws.cell(r, COL["cum"]).value),
        })

    anchor = None
    for s in steps:
        if str(s["actual"]).strip().lower() == "t=0":
            anchor = s
            break
    if anchor is None:
        for s in steps:
            if s["cum"] == timedelta(0) and "gas" in s["what"].lower():
                anchor = s
                break
    if anchor is None:
        LOG.warning("No t=0 anchor in %s; writing placeholder", xlsx.name)
        return _write_placeholder_injection(out_csv)

    # Pure protocol-based anchor (reference clock): the t=0 step's "target" time
    # on the experiment date. NO EXIF/image reads.
    exp_date = experiment_date or _date_from_folder(injection_folder, xlsx)
    anchor_dt = _coerce_time_on_date(anchor["target"], exp_date)
    if anchor_dt is None:
        LOG.warning("Cannot resolve anchor datetime in %s; placeholder", xlsx.name)
        return _write_placeholder_injection(out_csv)

    port = "L1"
    rows = []
    rate_steps = [s for s in steps if s["cum"] is not None and s["flow"] is not None]
    for i, s in enumerate(rate_steps):
        todo = s["todo"].lower()
        if "open l1" in todo:
            port = "L1"
        elif "open l2" in todo:
            port = "L2"
        start_dt = anchor_dt + s["cum"]
        if i + 1 < len(rate_steps):
            end_dt = anchor_dt + rate_steps[i + 1]["cum"]
        else:
            secs = float(s["seconds"]) if s["seconds"] else 0.0
            end_dt = start_dt + timedelta(seconds=secs)
        try:
            flow_pct = float(s["flow"])
        except (TypeError, ValueError):
            continue
        loc = PORT_LOCATION.get(port, (0.0, 0.0))
        rows.append({
            "location_x": loc[0], "location_y": loc[1],
            "start": start_dt, "end": end_dt,
            "flow_percent": flow_pct,
            "rate_sccm": round(flow_pct / 100.0 * nominal_sccm, 6),
        })

    out_csv.parent.mkdir(parents=True, exist_ok=True)
    with out_csv.open("w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(["id", "location_x", "location_y", "start", "end",
                    "flow_percent", "rate_sccm", "density kg/m3"])
        for idx, r in enumerate(rows, start=1):
            w.writerow([idx, r["location_x"], r["location_y"],
                        r["start"].isoformat(sep=" "), r["end"].isoformat(sep=" "),
                        r["flow_percent"], r["rate_sccm"], DENSITY_KG_M3])
    LOG.info("Wrote %s (%d intervals, anchor target=%s @ %s)",
             out_csv, len(rows), anchor.get("target"), anchor_dt)
    return len(rows)


def _write_placeholder_injection(out_csv: Path) -> int:
    out_csv.parent.mkdir(parents=True, exist_ok=True)
    with out_csv.open("w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(["id", "location_x", "location_y", "start", "end",
                    "flow_percent", "rate_sccm", "density kg/m3"])
    LOG.warning("Wrote empty placeholder %s (needs manual review)", out_csv)
    return 0


def _date_from_folder(folder: Optional[Path], xlsx: Optional[Path] = None) -> Optional[datetime]:
    if folder:
        m = FOLDER_RE.search(folder.name)
        if m:
            try:
                return datetime.strptime(m["date"], "%y%m%d")
            except ValueError:
                pass  # e.g. AC22 typo (month 15) -> fall through
    return None


def _xlsx_header_dates(xlsx: Path) -> list:
    """All date/datetime cells in the protocol header block (rows 1-6)."""
    import openpyxl
    out = []
    try:
        wb = openpyxl.load_workbook(xlsx, data_only=True)
        ws = wb[wb.sheetnames[0]]
        for r in range(1, 7):
            for c in range(1, 27):
                v = ws.cell(r, c).value
                if isinstance(v, datetime):
                    out.append(v)
                elif hasattr(v, "year") and hasattr(v, "month") and hasattr(v, "day"):
                    out.append(datetime(v.year, v.month, v.day))
    except Exception:  # noqa: BLE001
        pass
    return out


def _date_from_xlsx(xlsx: Optional[Path], hint_token: Optional[str] = None) -> Optional[datetime]:
    """Experiment date from the protocol: filename token first, else an in-sheet
    header date. When the folder date is a typo, ``hint_token`` (the malformed
    folder digits) is used to pick the header date whose %y%m%d digits are the
    same multiset (i.e. a transposition like AC22 221511 <-> 221115)."""
    if not xlsx:
        return None
    m = FOLDER_RE.search(xlsx.name)
    if m:
        try:
            return datetime.strptime(m["date"], "%y%m%d")
        except ValueError:
            pass
    dates = _xlsx_header_dates(xlsx)
    if not dates:
        return None
    if hint_token:
        from collections import Counter
        want = Counter(hint_token)
        for d in dates:
            if Counter(d.strftime("%y%m%d")) == want:
                return datetime(d.year, d.month, d.day)
    return datetime(dates[0].year, dates[0].month, dates[0].day)


def resolve_experiment_date(inj_folders: list, xlsx: Optional[Path]) -> Optional[datetime]:
    """Pick the experiment date once, preferring a valid folder-name date, then a
    protocol date (filename or in-sheet header). Pure protocol data - no EXIF."""
    hint = None
    for f in (inj_folders or []):
        d = _date_from_folder(f)
        if d:
            return d
        if hint is None:
            m = FOLDER_RE.search(f.name)
            if m:
                hint = m["date"]
    return _date_from_xlsx(xlsx, hint_token=hint)


def _coerce_time_on_date(time_val, base_date: Optional[datetime]) -> Optional[datetime]:
    if time_val is None or base_date is None:
        return None
    if isinstance(time_val, datetime):
        return time_val
    s = str(time_val).strip()
    m = re.match(r"(\d{1,2}):(\d{2}):(\d{2})", s)
    if not m:
        return None
    return base_date.replace(hour=int(m[1]), minute=int(m[2]), second=int(m[3]))


# =========================================================================
# Driver
# =========================================================================
def _is_injection(name: str) -> bool:
    return "injection" in name.lower()


def _is_resting(name: str) -> bool:
    n = name.lower()
    return "resting" in n or re.search(r"(^|_)rest(_|ing|\d|$)", n) is not None


def _folder_min_dsc(folder: Path) -> int:
    nums = [int(m[1]) for q in folder.glob("*.JPG")
            for m in [DSC_RE.search(q.stem)] if m]
    return min(nums) if nums else 10**9


def list_phase_folders(exp_dir: Path):
    """Return (injection_folders, resting_folders), ordered chronologically."""
    inj, rest = [], []
    for sub in exp_dir.iterdir():
        if not sub.is_dir() or not FOLDER_RE.search(sub.name):
            continue
        if _is_injection(sub.name):
            inj.append(sub)
        elif _is_resting(sub.name):
            rest.append(sub)
    inj.sort(key=_folder_min_dsc)
    rest.sort(key=_folder_min_dsc)
    return inj, rest


def imaging_csv_name(tag: str, k: int, n_total: int) -> str:
    suffix = f"{tag}{k}" if n_total > 1 else tag
    return f"imaging_protocol_{suffix}.csv"


def find_protocol_xlsx(exp_dir: Path) -> Optional[Path]:
    exp_key = exp_dir.name.lower().replace(" ", "")
    cands = [p for p in exp_dir.glob("*.xlsx")
             if not p.name.startswith("~$") and "overflow" not in p.name.lower()]
    # 1) prefer a file explicitly named like a protocol (the standard table);
    #    this beats e.g. AC18_overflow.xlsx, which has no scripted-steps table.
    for p in cands:
        if "protocol" in p.name.lower():
            return p
    # 2) else a file named like the experiment (e.g. AC32/AC32.xlsx).
    for p in cands:
        if exp_key in p.name.lower().replace(" ", ""):
            return p
    # 3) fallback one level up: AC60/AC61 keep "Albus protocol NN.xlsx" at root.
    m = re.search(r"AC(\d+)", exp_dir.name, re.IGNORECASE)
    if m:
        num = m.group(1)
        for p in exp_dir.parent.glob("*.xlsx"):
            n = p.name.lower()
            if "overflow" in n or p.name.startswith("~$"):
                continue
            if "protocol" in n and re.search(rf"\b0*{num}\b", n):
                return p
    return None


def process_experiment(exp_dir: Path, out_root: Path) -> dict:
    out_dir = out_root / exp_dir.name.lower()
    inj_folders, rest_folders = list_phase_folders(exp_dir)
    summary = {"experiment": exp_dir.name, "phases": {}, "injection": 0}

    xlsx = find_protocol_xlsx(exp_dir)
    experiment_date = resolve_experiment_date(inj_folders + rest_folders, xlsx)

    def emit(folders: list, tag: str) -> None:
        for k, f in enumerate(folders):
            out_csv = out_dir / imaging_csv_name(tag, k, len(folders))
            summary["phases"][f.name] = build_imaging_protocol(
                f, out_csv, experiment_date=experiment_date)

    emit(inj_folders, "inj")
    emit(rest_folders, "rest")

    anchor_folder = inj_folders[0] if inj_folders else None
    if xlsx:
        summary["injection"] = build_injection_protocol(
            xlsx, anchor_folder, out_dir / "injection_protocol.csv",
            experiment_date=experiment_date)
    else:
        LOG.warning("No protocol xlsx for %s", exp_dir.name)
    if not inj_folders:
        LOG.warning("%s: no injection folder found", exp_dir.name)
    if not rest_folders:
        LOG.warning("%s: no resting folder found", exp_dir.name)
    return summary


def main() -> None:
    # CLI entry: pure protocol-based timing (no EXIF/image reads).
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--albus-root", required=True, type=Path)
    ap.add_argument("--out", required=True, type=Path, help="output protocols root")
    ap.add_argument("--experiments", nargs="*", default=[], help="e.g. AC53 AC60")
    ap.add_argument("--all", action="store_true", help="process every ACxx folder")
    args = ap.parse_args()
    logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")

    root = args.albus_root
    if args.all:
        exps = sorted([d for d in root.iterdir()
                       if d.is_dir() and re.fullmatch(r"AC\d+", d.name)],
                      key=lambda d: int(d.name[2:]))
    else:
        exps = [root / e for e in args.experiments]

    for exp in exps:
        if not exp.exists():
            LOG.error("Missing %s", exp)
            continue
        s = process_experiment(exp, args.out)
        LOG.info("DONE %s: imaging=%s injection=%d intervals",
                 s["experiment"], s["phases"], s["injection"])


if __name__ == "__main__":
    main()
